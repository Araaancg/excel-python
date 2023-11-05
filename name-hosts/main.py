import openpyxl as xl

# Load the Excel file
workbook = xl.load_workbook('excelEjemplo.xlsx')

# Access the sheets
email_sheet = workbook['emails']
host_sheet = workbook['hosts']

# Create a new sheet for the combined data
combined_sheet = workbook.create_sheet(title='combinedData')

# Extract usernames from email addresses (from column A, skipping the header)
usernames = [row[0].split('@')[0] for row in email_sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True)]

# Create a dictionary to store the mapping of usernames to hosts
user_host_mapping = {}

# Iterate through host rows (from column A, skipping the header)
for row in host_sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
    host = row[0]
    username = ''.join(filter(str.isalpha, host))
    if username in usernames:
        user_host_mapping[username] = host

# Write the email and corresponding host to the combined sheet
for username in usernames:
    email = f'{username}@email.com'
    host = user_host_mapping.get(username, 'No matching host found')
    combined_sheet.append([email, host])

# Save the workbook
workbook.save('excelEjemplo.xlsx')
